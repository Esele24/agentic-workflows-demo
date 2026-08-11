"""Merge scraped DailyRemote jobs into a master Excel workbook with dedupe.

Usage:
    python tools/update_job_workbook.py --in tmp/social_media_raw.json --workbook output/dailyremote_social_media.xlsx

Behavior:
- Dedupe key is job_id.
- New jobs get first_seen = today and is_new = TRUE.
- Jobs seen again get last_seen = today and is_new = FALSE.
- Jobs that vanished from the site keep their row with active = FALSE.
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

TRACKING_COLUMNS = ["first_seen", "last_seen", "is_new", "active"]
COLUMN_ORDER = [
    "is_new", "active", "title", "company", "job_type", "location", "experience",
    "salary_raw", "salary_min", "salary_max", "salary_currency", "salary_period",
    "category", "tags", "posted_date_est", "posted_relative", "first_seen",
    "last_seen", "snippet", "url", "search_term", "job_id",
]
COLUMN_WIDTHS = {
    "is_new": 8, "active": 8, "title": 52, "company": 22, "job_type": 12,
    "location": 24, "experience": 13, "salary_raw": 24, "salary_min": 12,
    "salary_max": 12, "salary_currency": 10, "salary_period": 12,
    "category": 14, "tags": 28,
    "posted_date_est": 15, "posted_relative": 15, "first_seen": 12,
    "last_seen": 12, "snippet": 80, "url": 45, "search_term": 16, "job_id": 10,
}


def load_scraped(path: Path) -> pd.DataFrame:
    records = json.loads(path.read_text(encoding="utf-8"))
    if not records:
        print("ERROR: input JSON contains no records", file=sys.stderr)
        sys.exit(1)
    df = pd.DataFrame(records)
    df["job_id"] = df["job_id"].astype(str)
    return df


def merge(scraped: pd.DataFrame, existing: pd.DataFrame | None) -> pd.DataFrame:
    today = date.today().isoformat()
    scraped = scraped.copy()
    scraped["first_seen"] = today
    scraped["last_seen"] = today
    scraped["is_new"] = True
    scraped["active"] = True

    if existing is None or existing.empty:
        return scraped

    existing = existing.copy()
    existing["job_id"] = existing["job_id"].astype(str)
    known = set(existing["job_id"])
    scraped_ids = set(scraped["job_id"])

    # refresh rows that are still live
    existing["is_new"] = False
    existing["active"] = existing["job_id"].isin(scraped_ids)
    existing.loc[existing["active"], "last_seen"] = today

    # update volatile fields (salary can appear later, snippet can change)
    refresh_cols = [
        c for c in scraped.columns
        if c not in TRACKING_COLUMNS + ["job_id", "posted_relative", "posted_date_est"]
    ]
    live = scraped.set_index("job_id")
    existing = existing.set_index("job_id")
    overlap = existing.index.intersection(live.index)
    existing.loc[overlap, refresh_cols] = live.loc[overlap, refresh_cols]
    existing = existing.reset_index()

    new_rows = scraped[~scraped["job_id"].isin(known)]
    return pd.concat([existing, new_rows], ignore_index=True)


def write_workbook(df: pd.DataFrame, path: Path) -> None:
    df = df.copy()
    for col in COLUMN_ORDER:
        if col not in df.columns:
            df[col] = ""
    df = df[COLUMN_ORDER]
    df = df.sort_values(
        by=["is_new", "active", "posted_date_est"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Jobs", index=False)
        sheet = writer.sheets["Jobs"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for idx, col in enumerate(COLUMN_ORDER, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS[col]
        url_col = COLUMN_ORDER.index("url") + 1
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=url_col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge scraped jobs into master Excel")
    parser.add_argument("--in", dest="input", required=True, help="Scraped JSON path")
    parser.add_argument("--workbook", required=True, help="Master .xlsx path")
    args = parser.parse_args()

    scraped = load_scraped(Path(args.input))
    workbook_path = Path(args.workbook)
    existing = None
    if workbook_path.exists():
        existing = pd.read_excel(workbook_path, sheet_name="Jobs", dtype={"job_id": str})

    merged = merge(scraped, existing)
    write_workbook(merged, workbook_path)

    total = len(merged)
    new = int(merged["is_new"].sum())
    inactive = int((~merged["active"].astype(bool)).sum())
    print(f"Workbook: {workbook_path}")
    print(f"Total jobs: {total} | new this run: {new} | no longer listed: {inactive}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
